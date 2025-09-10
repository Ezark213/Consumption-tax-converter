import pandas as pd
import openpyxl
from typing import Dict, List, Any
from .base import BaseParser

class MoneyforwardParser(BaseParser):
    """
    マネーフォワード会計の勘定科目別税区分集計表パーサー
    """
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls', '.pdf']
        self.parser_name = "moneyforward"
    
    def detect_format(self, file_path: str) -> bool:
        """
        マネーフォワード形式のExcel/PDFファイルかどうかを判定
        """
        if not self._validate_file(file_path):
            return False
        
        file_extension = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_extension == '.pdf':
                # PDFファイルの場合
                import PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    
                    # 最初の数ページからマネーフォワードの特徴的なテキストを検索
                    for page_num in range(min(3, len(pdf_reader.pages))):
                        page = pdf_reader.pages[page_num]
                        text = page.extract_text()
                        
                        # マネーフォワードの特徴的なキーワードをチェック
                        if 'マネーフォワード' in text or 'MoneyForward' in text:
                            return True
                        
                        # 勘定科目別税区分集計表（マネーフォワード形式）
                        if '勘定科目別税区分集計表' in text and 'Cognite' in text:
                            return True
                        
                        # インボイス列があるのもマネーフォワードの特徴
                        if 'インボイス' in text and '税区分' in text:
                            return True
                
                return False
                
            else:
                # Excelファイルの場合
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                
                # シート名やセル内容からマネーフォワードの特徴を検出
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # 最初の数行を確認
                    for row in range(1, min(10, sheet.max_row + 1)):
                        for col in range(1, min(10, sheet.max_column + 1)):
                            cell_value = sheet.cell(row, col).value
                            if cell_value:
                                cell_text = str(cell_value)
                                
                                # マネーフォワードの特徴的なキーワード
                                mf_keywords = [
                                    'マネーフォワード',
                                    'MoneyForward',
                                    '勘定科目別税区分集計表',
                                    '税区分集計'
                                ]
                                
                                for keyword in mf_keywords:
                                    if keyword in cell_text:
                                        workbook.close()
                                        return True
                
                workbook.close()
                return False
            
        except Exception:
            return False
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        マネーフォワード形式のExcelを解析
        """
        try:
            # Excelファイルを読み込み
            df = pd.read_excel(file_path, sheet_name=None)  # 全シートを読み込み
            
            sales_data = []
            purchase_data = []
            
            # 各シートを確認
            for sheet_name, sheet_df in df.items():
                # 売上データと仕入データを抽出
                sheet_sales, sheet_purchases = self._extract_data_from_sheet(sheet_df)
                sales_data.extend(sheet_sales)
                purchase_data.extend(sheet_purchases)
            
            # メタデータ抽出
            metadata = self._extract_metadata_from_excel(file_path)
            
            return self._create_standard_output(sales_data, purchase_data, metadata)
            
        except Exception as e:
            return {
                'sales_items': [],
                'purchase_items': [],
                'taxable_sales_total': 0,
                'taxable_purchases_total': 0,
                'parser_type': self.parser_name,
                'warnings': [],
                'errors': [f"Parse error: {str(e)}"]
            }
    
    def _extract_data_from_sheet(self, df: pd.DataFrame) -> tuple[List[Dict], List[Dict]]:
        """
        シートからデータを抽出
        """
        sales_data = []
        purchase_data = []
        
        # データフレームが空の場合は空リストを返す
        if df.empty:
            return sales_data, purchase_data
        
        # ヘッダー行を特定
        header_row = self._find_header_row(df)
        if header_row == -1:
            return sales_data, purchase_data
        
        # ヘッダー行以降のデータを処理
        data_df = df.iloc[header_row + 1:].copy()
        
        # 列名を設定
        data_df.columns = df.iloc[header_row].values
        
        for index, row in data_df.iterrows():
            account_name = self._standardize_account_name(row.get('勘定科目', ''))
            
            if not account_name:
                continue
            
            # 税区分別の金額を処理
            for col_name in data_df.columns:
                if '売上' in col_name and account_name:
                    amount = self._extract_numeric_value(row.get(col_name, 0))
                    if amount > 0:
                        tax_rate = self._extract_tax_rate_from_column(col_name)
                        sales_data.append({
                            'account_name': account_name,
                            'tax_rate': tax_rate,
                            'amount': amount,
                            'taxable_amount': amount if self._is_taxable(tax_rate) else 0
                        })
                
                elif '仕入' in col_name and account_name:
                    amount = self._extract_numeric_value(row.get(col_name, 0))
                    if amount > 0:
                        tax_rate = self._extract_tax_rate_from_column(col_name)
                        purchase_data.append({
                            'account_name': account_name,
                            'tax_rate': tax_rate,
                            'amount': amount,
                            'taxable_amount': amount if self._is_taxable(tax_rate) else 0
                        })
        
        return sales_data, purchase_data
    
    def _find_header_row(self, df: pd.DataFrame) -> int:
        """
        ヘッダー行を特定
        """
        for i, row in df.iterrows():
            row_text = ' '.join(str(cell) for cell in row.values if pd.notna(cell))
            
            # ヘッダーの特徴的なキーワード
            header_keywords = ['勘定科目', '税区分', '金額', '合計']
            
            keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
            if keyword_count >= 2:
                return i
        
        return -1
    
    def _extract_tax_rate_from_column(self, column_name: str) -> str:
        """
        列名から税率を抽出
        """
        import re
        
        # 税率パターンを検索
        tax_patterns = [
            r'(\d+)%',
            r'軽減(\d+)%',
            r'標準(\d+)%',
            r'非課税',
            r'不課税',
            r'輸出'
        ]
        
        for pattern in tax_patterns:
            match = re.search(pattern, column_name)
            if match:
                return match.group(0)
        
        return '不明'
    
    def _is_taxable(self, tax_rate: str) -> bool:
        """
        課税対象かどうかを判定
        """
        taxable_patterns = ['10%', '8%', '軽減8%', '標準10%']
        return any(pattern in tax_rate for pattern in taxable_patterns)
    
    def _extract_metadata_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Excelファイルからメタデータを抽出
        """
        metadata = {}
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            # 最初のシートからメタデータを抽出
            first_sheet = workbook[workbook.sheetnames[0]]
            
            # 期間や会社名の抽出
            for row in range(1, min(20, first_sheet.max_row + 1)):
                for col in range(1, min(10, first_sheet.max_column + 1)):
                    cell_value = first_sheet.cell(row, col).value
                    if cell_value:
                        cell_text = str(cell_value)
                        
                        # 期間の抽出
                        import re
                        period_pattern = r'(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})'
                        period_match = re.search(period_pattern, cell_text)
                        
                        if period_match and 'period_start' not in metadata:
                            metadata['period_extracted'] = cell_text
                        
                        # 会社名の抽出
                        if any(keyword in cell_text for keyword in ['株式会社', '有限会社', '合同会社']) and 'company_name' not in metadata:
                            metadata['company_name'] = cell_text
            
            workbook.close()
            
        except Exception:
            pass
        
        return metadata